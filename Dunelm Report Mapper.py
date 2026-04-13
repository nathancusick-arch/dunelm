import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.formula.translate import Translator
from openpyxl.cell.cell import MergedCell
from copy import copy
import io
from datetime import datetime

st.title("Dunelm Report Generator")

st.write("""
Upload:
1. Latest audits export
2. Last week's LIVE report

The tool will automatically generate:
- This week's LIVE report
- This week's completed report
""")

csv_file = st.file_uploader("Upload audits_basic_data_export.csv", type=["csv"])
live_file = st.file_uploader("Upload last week's LIVE report", type=["xlsx"])

COLUMN_MAP = {
    "Order": "order_internal_id",
    "Client": "client_name",
    "Visit": "internal_id",
    "Site": "site_internal_id",
    "Order Deadline": "responsibility",
    "Responsibility": "site_name",
    "Premises Name": "site_address_1",
    "Address1": "site_address_2",
    "Address2": "site_address_3",
    "Address3": None,
    "City": None,
    "Post Code": "site_post_code",
    "Submitted Date": "submitted_date",
    "Approved Date": "approval_date",
    "Item to order": "item_to_order",
    "Actual Visit Date": "date_of_visit",
    "Actual Visit Time": "time_of_visit",
    "AM / PM": None,
    "Pass-Fail": "primary_result",
    "Pass-Fail2": "secondary_result",
    "Abort Reason": "Please detail why you were unable to conduct this audit:",
    "Extra Site 1": "site_code",
    "Extra Site 2": None,
    "Extra Site 3": None,
    "Extra Site 4": None,
}

OUTPUT_COLUMNS = list(COLUMN_MAP.keys())

def map_data(df):
    def map_value(row, mapping):
        if mapping is None:
            return ""
        return str(row.get(mapping, "")).strip()

    final_df = pd.DataFrame(columns=OUTPUT_COLUMNS)

    for col in OUTPUT_COLUMNS:
        final_df[col] = df.apply(lambda r: map_value(r, COLUMN_MAP[col]), axis=1)

    for col in ["Actual Visit Date", "Submitted Date", "Approved Date"]:
        final_df[col] = pd.to_datetime(final_df[col], errors="coerce", dayfirst=True)

    final_df["Actual Visit Time"] = pd.to_datetime(
        final_df["Actual Visit Time"], errors="coerce"
    ).dt.time

    final_df["Extra Site 1"] = pd.to_numeric(final_df["Extra Site 1"], errors="coerce")

    av = final_df[~final_df["Item to order"].isin(["Allergens", "Restaurant", "On the Go"])].copy()
    narv = final_df[final_df["Item to order"].isin(["Allergens", "Restaurant", "On the Go"])].copy()

    return av, narv

def get_last_data_row(ws):
    for r in range(ws.max_row, 3, -1):
        if ws.cell(r, 1).value:
            return r
    return 4

def write(ws, df):
    for row in ws.iter_rows(min_row=4, max_col=25):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None

    for r, row in enumerate(df.itertuples(index=False), start=4):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(r, c, val)

            if c in [13, 14, 16]:
                if pd.notna(val):
                    cell.number_format = "DD/MM/YYYY"

            if c == 17 and pd.notna(val):
                cell.number_format = "HH:MM"

def append(ws, df):
    start = get_last_data_row(ws) + 1

    for r, row in enumerate(df.itertuples(index=False), start=start):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(r, c, val)

            if c in [13, 14, 16]:
                if pd.notna(val):
                    cell.number_format = "DD/MM/YYYY"

            if c == 17 and pd.notna(val):
                cell.number_format = "HH:MM"

def fix_formulas(ws):
    last_row = get_last_data_row(ws)

    for col in [27, 28]:
        col_letter = ws.cell(row=4, column=col).column_letter
        base = ws[f"{col_letter}4"].value

        for r in range(5, 1000):
            ws[f"{col_letter}{r}"] = None

        for r in range(5, last_row + 1):
            ws[f"{col_letter}{r}"] = Translator(
                base, origin=f"{col_letter}4"
            ).translate_formula(f"{col_letter}{r}")

def fix_summary(ws, start_row, data_ws):
    base_row = start_row + 1
    data_len = get_last_data_row(data_ws) - 3

    template = []
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(base_row, c)
        template.append({
            "value": cell.value,
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format
        })

    max_clear = ws.max_row + 200
    for r in range(base_row, max_clear):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

    for i in range(data_len):
        for c, t in enumerate(template, start=1):
            if t["value"]:
                cell = ws.cell(
                    base_row + i,
                    c,
                    Translator(t["value"], origin=f"A{base_row}")
                    .translate_formula(f"A{base_row+i}")
                )
                cell.font = t["font"]
                cell.fill = t["fill"]
                cell.border = t["border"]
                cell.alignment = t["alignment"]
                cell.number_format = t["number_format"]

    final_row = base_row + data_len - 1
    if ws.max_row > final_row:
        ws.delete_rows(final_row + 1, ws.max_row - final_row)

def trim_sheet(ws):
    last_row = get_last_data_row(ws)
    if ws.max_row > last_row:
        ws.delete_rows(last_row + 1, ws.max_row - last_row)

if csv_file and live_file:

    df = pd.read_csv(csv_file, dtype=str).fillna("")
    av_df, narv_df = map_data(df)

    wb = load_workbook(live_file)

    weekly = wb["Weekly"]
    narv_weekly = wb["NARV Weekly"]
    ytd = wb["YTD"]
    narv_ytd = wb["NARV YTD"]

    existing = {
        str(r[2].value)
        for ws in [weekly, narv_weekly]
        for r in ws.iter_rows(min_row=4, max_col=3)
        if r[2].value
    }

    av_df = av_df[~av_df["Visit"].isin(existing)]
    narv_df = narv_df[~narv_df["Visit"].isin(existing)]

    write(weekly, av_df)
    write(narv_weekly, narv_df)

    append(ytd, av_df)
    append(narv_ytd, narv_df)

    for ws in [weekly, narv_weekly, ytd, narv_ytd]:
        fix_formulas(ws)

    fix_summary(wb["Weekly Summary Table"], 21, weekly)
    fix_summary(wb["Allergens Weekly Summary Table"], 16, narv_weekly)

    for ws in [weekly, narv_weekly, ytd, narv_ytd]:
        trim_sheet(ws)

    today = datetime.today().strftime("%d.%m.%Y")

    live_buffer = io.BytesIO()
    wb.save(live_buffer)
    live_buffer.seek(0)

    # ============================================================
    # COMPLETED REPORT
    # ============================================================

    wb_completed = load_workbook(live_buffer)

    weekly = wb_completed["Weekly"]
    narv_weekly = wb_completed["NARV Weekly"]
    regions = wb_completed["Regions"]
    ytd = wb_completed["YTD"]
    narv_ytd = wb_completed["NARV YTD"]

    ws_summary = wb_completed["Weekly Summary Table"]
    ws_narv_summary = wb_completed["Allergens Weekly Summary Table"]
    ws_area = wb_completed[" Performance by Area"]
    ws_narv_area = wb_completed[" Allergens Performance by Area"]

    lookup = {}
    for row in regions.iter_rows(min_row=2):
        if row[0].value:
            lookup[str(row[0].value)] = (row[1].value, row[2].value, row[3].value)

    def extract(ws):
        data = []
        for r in range(4, ws.max_row + 1):
            code = ws.cell(r, 22).value
            if not code:
                continue

            code = int(code)
            name, am, pot = lookup.get(str(code), (None, None, None))

            data.append({
                "Code": code,
                "Name": name,
                "Pot": pot,
                "AM": am,
                "Product": ws.cell(r, 15).value,
                "Date": ws.cell(r, 16).value,
                "Result": str(ws.cell(r, 19).value).lower()
            })

        return pd.DataFrame(data)

    weekly_df = extract(weekly)
    narv_df = extract(narv_weekly)
    ytd_df = extract(ytd)
    narv_ytd_df = extract(narv_ytd)

    def summary(df):
        valid = df[df["Result"].isin(["pass", "fail"])]
        total = len(valid)
        p = (valid["Result"] == "pass").sum()
        f = (valid["Result"] == "fail").sum()
        return p, f, total, (p/total if total else 0), (f/total if total else 0)

    wp, wf, wt, wpct, wfct = summary(weekly_df)
    yp, yf, yt, ypct, yfct = summary(ytd_df)

    ws_summary["B6"], ws_summary["C6"] = wp, wpct
    ws_summary["B7"], ws_summary["C7"] = wf, wfct
    ws_summary["B9"] = wt

    ws_summary["D6"], ws_summary["E6"] = yp, ypct
    ws_summary["D7"], ws_summary["E7"] = yf, yfct
    ws_summary["D9"] = yt

    products = ["Knife", "Knife - No ID", "Click & Collect"]

    def product_stats(df, product):
        sub = df[df["Product"] == product]
        valid = sub[sub["Result"].isin(["pass", "fail"])]
        total = len(valid)
        p = (valid["Result"] == "pass").sum()
        return total, (p/total if total else None)

    for i, prod in enumerate(products, start=14):
        total, pct = product_stats(weekly_df, prod)
        ws_summary[f"B{i}"] = total
        ws_summary[f"C{i}"] = pct if pct is not None else "-"

        total, pct = product_stats(ytd_df, prod)
        ws_summary[f"D{i}"] = total
        ws_summary[f"E{i}"] = pct if pct is not None else "-"

    ws_summary["B18"] = sum(ws_summary[f"B{i}"].value or 0 for i in range(14,17))
    ws_summary["D18"] = sum(ws_summary[f"D{i}"].value or 0 for i in range(14,17))

    wp, wf, wt, wpct, wfct = summary(narv_df)
    yp, yf, yt, ypct, yfct = summary(narv_ytd_df)

    ws_narv_summary["B6"], ws_narv_summary["C6"] = wp, wpct
    ws_narv_summary["B7"], ws_narv_summary["C7"] = wf, wfct
    ws_narv_summary["B9"] = wt

    ws_narv_summary["D6"], ws_narv_summary["E6"] = yp, ypct
    ws_narv_summary["D7"], ws_narv_summary["E7"] = yf, yfct
    ws_narv_summary["D9"] = yt

    restaurant = narv_df[narv_df["Product"] == "Restaurant"]
    valid = restaurant[restaurant["Result"].isin(["pass", "fail"])]

    if len(valid):
        ws_narv_summary["B13"] = (valid["Result"] == "pass").sum() / len(valid)
    else:
        ws_narv_summary["B13"] = "-"

    def write_table(ws, df, start_row):
        for i, row in df.iterrows():
            r = start_row + i
            ws.cell(r, 1, row["Code"])
            ws.cell(r, 2, row["Name"])
            ws.cell(r, 3, row["Pot"])
            ws.cell(r, 4, row["AM"])
            ws.cell(r, 5, row["Product"])
            ws.cell(r, 6, row["Date"])
            ws.cell(r, 7, row["Result"])

    write_table(ws_summary, weekly_df, 22)
    write_table(ws_narv_summary, narv_df, 17)

    # ============================================================
    # PERFORMANCE BY AREA
    # ============================================================

    def area_table(ws, weekly_df, ytd_df):
        start_row = 8

        ams = []
        r = start_row
        while ws.cell(r, 1).value:
            ams.append(ws.cell(r, 1).value)
            r += 1

        for i, am in enumerate(ams):
            row = start_row + i

            w = weekly_df[weekly_df["AM"] == am]
            valid_w = w[w["Result"].isin(["pass", "fail"])]

            comp = len(valid_w)
            fail = (valid_w["Result"] == "fail").sum()
            pas = (valid_w["Result"] == "pass").sum()

            ws[f"B{row}"] = comp
            ws[f"C{row}"] = fail
            ws[f"D{row}"] = pas
            ws[f"E{row}"] = (pas / comp) if comp else "-"

            y = ytd_df[ytd_df["AM"] == am]
            valid_y = y[y["Result"].isin(["pass", "fail"])]

            comp = len(valid_y)
            fail = (valid_y["Result"] == "fail").sum()
            pas = (valid_y["Result"] == "pass").sum()

            ws[f"G{row}"] = comp
            ws[f"H{row}"] = fail
            ws[f"I{row}"] = pas
            ws[f"J{row}"] = (pas / comp) if comp else "-"

        total_row = start_row + len(ams) + 1

        weekly_valid = weekly_df[weekly_df["Result"].isin(["pass", "fail"])]
        ytd_valid = ytd_df[ytd_df["Result"].isin(["pass", "fail"])]

        ws[f"B{total_row}"] = len(weekly_valid)
        ws[f"C{total_row}"] = (weekly_valid["Result"] == "fail").sum()
        ws[f"D{total_row}"] = (weekly_valid["Result"] == "pass").sum()
        ws[f"E{total_row}"] = (
            (weekly_valid["Result"] == "pass").sum() / len(weekly_valid)
            if len(weekly_valid) else "-"
        )

        ws[f"G{total_row}"] = len(ytd_valid)
        ws[f"H{total_row}"] = (ytd_valid["Result"] == "fail").sum()
        ws[f"I{total_row}"] = (ytd_valid["Result"] == "pass").sum()
        ws[f"J{total_row}"] = (
            (ytd_valid["Result"] == "pass").sum() / len(ytd_valid)
            if len(ytd_valid) else "-"
        )

    area_table(ws_area, weekly_df, ytd_df)
    area_table(ws_narv_area, narv_df, narv_ytd_df)

    keep = [
        "Weekly Summary Table",
        " Performance by Area",
        "Allergens Weekly Summary Table",
        " Allergens Performance by Area"
    ]

    for sheet in list(wb_completed.sheetnames):
        if sheet not in keep:
            del wb_completed[sheet]

    completed_buffer = io.BytesIO()
    wb_completed.active = wb_completed.sheetnames.index("Weekly Summary Table")
    wb_completed.save(completed_buffer)
    completed_buffer.seek(0)

    st.success("Reports generated successfully!")

    st.download_button("Download LIVE Report", live_buffer,
                       file_name=f"Weekly Report format - {today} LIVE.xlsx")

    st.download_button("Download Completed Report", completed_buffer,
                       file_name=f"Weekly Report format - {today}.xlsx")

    # ============================================================
    # EMAIL TEXT
    # ============================================================

    # AV
    av_p, av_f, av_total, av_pct, _ = summary(weekly_df)
    av_pass_rate = f"{round(av_pct * 100)}%"
    av_completed = av_total

    # Allergens
    narv_p, narv_f, narv_total, narv_pct, _ = summary(narv_df)
    narv_pass_rate = f"{round(narv_pct * 100)}%"
    narv_completed = narv_total

    email_text = (
        f"All,\n\n"
        f"Please find attached the Serve Legal weekly report. As you’ll see, "
        f"the age-verification pass rate was {av_pass_rate} based on {av_completed} completed audits, "
        f"and the allergens pass rate was {narv_pass_rate} based on {narv_completed} completed audits.\n\n"
        f"I hope you find the attached useful."
    )

    st.markdown("### Email Text")
    st.code(email_text, language="text")
